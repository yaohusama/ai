# coding=utf-8
import pandas as pd
import xlrd
import openpyxl
import datetime
import re
mid_date=datetime.date(1900,1,10)
import numpy as np
def transform_date(cell):
    cell_tmp = str(cell.value)
    if re.match('Detection',cell_tmp):
        return mid_date
    if re.match('<Null>',cell_tmp):#筛选空值
        return mid_date
    if len(cell_tmp.split('/'))!=1:#如果是另一种日期格式再处理
        return mid_date
    else :
        cell_tmp = str(cell_tmp).split(' ')[0]
        date_tmp = (cell_tmp.split('-'))

        date_tmp = [(int)(i) for i in date_tmp]
        cell_tmp = datetime.date(date_tmp[0], date_tmp[1], date_tmp[2])
        return cell_tmp
wb = openpyxl.load_workbook('2021MCMProblemC_DataSet.xlsx')
sheets = wb.sheetnames
sheet3 = wb['Sheet1']
ws = wb.active  # 当前活跃的表单
min_val= datetime.date(2022,12,12)
max_val= datetime.date(2000,12,12)
for cell in ws['B']:
    # print(cell.value)
    cell_tmp=cell.value
    # cell_tmp = datetime.datetime.now().strftime("%Y-%m-%d")
    cell_tmp=str(cell_tmp).split(' ')[0]
    # print(cell_tmp)
    if re.match('Detection',cell_tmp):
        continue
    if re.match('<Null>',cell_tmp):#筛选空值
        continue
    if len(cell_tmp.split('/'))!=1:#如果是另一种日期格式再处理
        continue
    else :
        date_tmp=(cell_tmp.split('-'))

        date_tmp=[(int)(i) for i in date_tmp]
        cell_tmp = datetime.date(date_tmp[0],date_tmp[1],date_tmp[2])

    if min_val >cell_tmp:
        min_val=cell_tmp
    if max_val<cell_tmp:
        max_val=cell_tmp
rng = pd.date_range(min_val,max_val)
index_list=[]
count_list=[]
for tmp in rng:
    print(1)
    count=0;
    for tmp1 in ws['B']:
        if transform_date(tmp1)==tmp:
            count=count+1
    if count!=0:
        index_list.append(tmp)
        count_list.append(count)

data=pd.DataFrame(count_list,index=index_list)
data.to_csv('时间分布后的记录.csv')
print(data)
#第二种方法，使用filter函数，然而并没有快多少，都没有秒出来是
# rng = pd.date_range(min_val,max_val)
# index_list=[]
# count_list=[]
# for i in rng:
#     test=filter(lambda x:transform_date(x)==i,list(ws['B'])) #要求字符串或者字节
#     if(len(list(test))!=0):
#         index_list.append(i)
#         count_list.append(len(list(test)))
# data=pd.DataFrame(count_list,index=index_list)
# data.to_csv('时间分布后的记录1.csv')
# print(data)

#常用的openpyxl库用法：
# print(ws)
# print(ws['A1'])  # 获取A列的第一个对象
# print(ws['A1'].value)
#
# c = ws['B1']
# print('Row {}, Column {} is {}'.format(c.row, c.column, c.value))  # 打印这个单元格对象所在的行列的数值和内容
# print('Cell {} is {}\n'.format(c.coordinate, c.value))  # 获取单元格对象的所在列的行数和值
# print(ws.cell(row=1, column=2)) # 获取第一行第二列的单元格
# print(ws.cell(row=1, column=2).value)
# for i in range(1, 8, 2): #  获取1,3,4,7 行第二列的值
#     print(i, ws.cell(row=i, column=2).value)
# colC = ws['C'] # 获取整个C列
# print(colC)
# row6 = ws[6]   # 获取第6行
# print(row6,type(row6))
# col_range = ws['B:C']
# row_range = ws[2:6]
#
# for col in col_range:  # 打印BC两列单元格中的值内容
#     for cell in col:
#         print(cell.value)
#
# for row in row_range:  # 打印 2-5行中所有单元格中的值
#     for cell in row:
#         print(cell.value)
#
# for row in ws.iter_rows(min_row=1, max_row=2, max_col=2):  # 打印1-2行，1-2列中的内容
#     for cell in row:
#         print(cell.value)
# print('{}行 {}列'.format(ws.max_row, ws.max_column))
# 把数字转换成字母
# print(wb.get_column_letter(2))

# 把字母转换成数字
# print(column_index_from_string('AAH'))
# from openpyxl.utils import get_column_letter, column_index_from_string
#
# # 根据列的数字返回字母
# print(get_column_letter(2))  # B
# # 根据字母返回列的数字
# print(column_index_from_string('D'))  # 4